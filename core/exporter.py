"""
闲鱼数据调研工具 - 数据导出模块 v2.0
修复：None值处理、异常容错
"""

import os
from datetime import datetime
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class Exporter:
    """数据导出器 v2.0"""

    def __init__(self, db):
        self.db = db

    def _safe_str(self, val):
        """安全转字符串，处理 None"""
        if val is None:
            return ""
        return str(val)

    def _safe_float(self, val):
        """安全转浮点数"""
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    def _safe_int(self, val):
        """安全转整数"""
        if val is None:
            return 0
        try:
            return int(val)
        except (ValueError, TypeError):
            return 0

    def export_to_excel(self, task_id=None, keyword="", output_path=None):
        """导出商品数据到 Excel"""
        items = self.db.get_items(task_id=task_id, limit=5000)
        if not items:
            raise ValueError("没有可导出的数据，请先采集商品")

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_keyword = re.sub(r'[\\/:*?"<>|]', '_', keyword or "all")[:20]
            filename = f"闲鱼数据_{safe_keyword}_{timestamp}.xlsx"
            output_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "exports", filename
            )

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        try:
            wb = Workbook()

            # ===== Sheet 1: 商品数据 =====
            ws_data = wb.active
            ws_data.title = "商品数据"

            headers = [
                "序号", "标题", "描述", "价格(元)", "原价(元)",
                "所在地", "卖家", "卖家等级",
                "浏览量", "想要数", "链接", "采集时间"
            ]
            self._write_headers(ws_data, headers)

            for i, item in enumerate(items, 1):
                try:
                    row = [
                        i,
                        self._safe_str(item.get("title", "")),
                        self._safe_str(item.get("description", "")),
                        self._safe_float(item.get("price", 0)),
                        self._safe_float(item.get("original_price", 0)),
                        self._safe_str(item.get("location", "")),
                        self._safe_str(item.get("seller_name", "")),
                        self._safe_str(item.get("seller_level", "")),
                        self._safe_int(item.get("views", 0)),
                        self._safe_int(item.get("wants", 0)),
                        self._safe_str(item.get("item_url", "")),
                        self._safe_str(item.get("collected_at", "")),
                    ]
                    ws_data.append(row)
                except Exception:
                    continue

            col_widths = [6, 40, 50, 10, 10, 12, 15, 12, 10, 10, 45, 20]
            for col_idx, width in enumerate(col_widths, 1):
                ws_data.column_dimensions[get_column_letter(col_idx)].width = width
            ws_data.freeze_panes = "A2"

            # ===== Sheet 2: 统计分析 =====
            ws_stats = wb.create_sheet("统计分析")

            stats = self.db.get_price_stats(task_id)
            self._write_section_header(ws_stats, 1, "价格统计")
            stats_data = [
                ("商品总数", self._safe_int(stats.get("count", 0))),
                ("平均价格", f"¥{self._safe_float(stats.get('avg_price', 0)):.2f}"),
                ("最低价格", f"¥{self._safe_float(stats.get('min_price', 0)):.2f}"),
                ("最高价格", f"¥{self._safe_float(stats.get('max_price', 0)):.2f}"),
            ]
            for i, (label, value) in enumerate(stats_data, 2):
                ws_stats.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws_stats.cell(row=i, column=2, value=value)

            # 标题分析
            try:
                from .analyzer import Analyzer
                analyzer = Analyzer(self.db)
                title_analysis = analyzer.analyze_titles(task_id)

                row_offset = 8
                self._write_section_header(ws_stats, row_offset, "标题文案分析")
                if "error" not in title_analysis:
                    ws_stats.cell(row=row_offset+1, column=1, value="平均标题长度").font = Font(bold=True)
                    ws_stats.cell(row=row_offset+1, column=2, value=f"{title_analysis.get('avg_title_length', 0)} 字")
                    ws_stats.cell(row=row_offset+2, column=1, value="标题数").font = Font(bold=True)
                    ws_stats.cell(row=row_offset+2, column=2, value=title_analysis.get("title_count", 0))

                    ws_stats.cell(row=row_offset+4, column=1, value="标题高频词 TOP 20").font = Font(bold=True, size=13)
                    ws_stats.cell(row=row_offset+5, column=1, value="关键词").font = Font(bold=True)
                    ws_stats.cell(row=row_offset+5, column=2, value="出现次数").font = Font(bold=True)
                    for i, w in enumerate(title_analysis.get("words", [])[:20]):
                        ws_stats.cell(row=row_offset+6+i, column=1, value=self._safe_str(w.get("word", "")))
                        ws_stats.cell(row=row_offset+6+i, column=2, value=self._safe_int(w.get("count", 0)))
            except Exception:
                pass

            # ===== Sheet 3: 文案汇总 =====
            ws_copy = wb.create_sheet("文案汇总")
            copy_headers = ["序号", "标题", "描述", "标题长度", "描述长度"]
            self._write_headers(ws_copy, copy_headers)

            for i, item in enumerate(items, 1):
                try:
                    title = self._safe_str(item.get("title", ""))
                    desc = self._safe_str(item.get("description", ""))
                    ws_copy.append([i, title, desc, len(title), len(desc)])
                except Exception:
                    continue

            ws_copy.column_dimensions["A"].width = 6
            ws_copy.column_dimensions["B"].width = 45
            ws_copy.column_dimensions["C"].width = 60
            ws_copy.column_dimensions["D"].width = 10
            ws_copy.column_dimensions["E"].width = 10

            wb.save(output_path)
            return output_path

        except Exception as e:
            raise Exception(f"导出Excel失败: {e}")

    def export_to_csv(self, task_id=None, keyword="", output_path=None):
        """导出为 CSV"""
        import csv

        items = self.db.get_items(task_id=task_id, limit=5000)
        if not items:
            raise ValueError("没有可导出的数据")

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_keyword = re.sub(r'[\\/:*?"<>|]', '_', keyword or "all")[:20]
            filename = f"闲鱼数据_{safe_keyword}_{timestamp}.csv"
            output_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "exports", filename
            )

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "序号", "标题", "描述", "价格", "原价", "所在地",
                "卖家", "浏览量", "想要数", "链接", "采集时间"
            ])
            for i, item in enumerate(items, 1):
                writer.writerow([
                    i,
                    self._safe_str(item.get("title", "")),
                    self._safe_str(item.get("description", "")),
                    self._safe_float(item.get("price", 0)),
                    self._safe_float(item.get("original_price", 0)),
                    self._safe_str(item.get("location", "")),
                    self._safe_str(item.get("seller_name", "")),
                    self._safe_int(item.get("views", 0)),
                    self._safe_int(item.get("wants", 0)),
                    self._safe_str(item.get("item_url", "")),
                    self._safe_str(item.get("collected_at", "")),
                ])

        return output_path

    def _write_headers(self, ws, headers):
        """写入表头"""
        header_fill = PatternFill(start_color="F5A623", end_color="F5A623", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _write_section_header(self, ws, row, title):
        """写入分区标题"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=14, color="F5A623")
